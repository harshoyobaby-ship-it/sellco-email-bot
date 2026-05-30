#!/usr/bin/env python3
"""Bulk email sender — one run, many recipients."""

from __future__ import annotations

import argparse
import csv
import io
import json
import logging
import smtplib
import ssl
import sys
import time
from dataclasses import dataclass
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from string import Template
from typing import Callable, Iterable

from dotenv import load_dotenv
import os

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("email_bot")


@dataclass
class SmtpConfig:
    host: str
    port: int
    username: str
    password: str
    use_tls: bool = True
    from_email: str | None = None
    from_name: str | None = None

    @classmethod
    def from_env(cls) -> SmtpConfig:
        host = os.getenv("SMTP_HOST", "").strip()
        port = int(os.getenv("SMTP_PORT", "587"))
        username = os.getenv("SMTP_USERNAME", "").strip()
        password = os.getenv("SMTP_PASSWORD", "").strip()
        use_tls = os.getenv("SMTP_USE_TLS", "true").lower() in ("1", "true", "yes")
        from_email = os.getenv("SMTP_FROM_EMAIL", username).strip() or username
        from_name = os.getenv("SMTP_FROM_NAME", "").strip() or None

        missing = [k for k, v in {
            "SMTP_HOST": host,
            "SMTP_USERNAME": username,
            "SMTP_PASSWORD": password,
        }.items() if not v]
        if missing:
            raise ValueError(f"Missing required env vars: {', '.join(missing)}")

        return cls(
            host=host,
            port=port,
            username=username,
            password=password,
            use_tls=use_tls,
            from_email=from_email,
            from_name=from_name,
        )

    @classmethod
    def from_mapping(cls, data: dict) -> SmtpConfig:
        username = str(data.get("username", "")).strip()
        from_email = str(data.get("from_email", username)).strip() or username
        use_tls = str(data.get("use_tls", "true")).lower() in ("1", "true", "yes")
        return cls(
            host=str(data.get("host", "")).strip(),
            port=int(data.get("port", 587)),
            username=username,
            password=str(data.get("password", "")).strip(),
            use_tls=use_tls,
            from_email=from_email,
            from_name=str(data.get("from_name", "")).strip() or None,
        )


@dataclass
class Recipient:
    email: str
    name: str = ""
    extra: dict[str, str] | None = None


def _cell_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _rows_to_recipients(rows: list[dict[str, str]]) -> list[Recipient]:
    if not rows:
        raise ValueError("Recipient list is empty")

    normalized_rows: list[dict[str, str]] = []
    for row in rows:
        normalized_rows.append(
            {_normalize_header(k): _cell_value(v) for k, v in row.items() if k}
        )

    if "email" not in normalized_rows[0]:
        raise ValueError("File must include an 'email' column")

    recipients: list[Recipient] = []
    for row in normalized_rows:
        email = row.get("email", "")
        if not email:
            continue
        name = row.get("name", "")
        extra = {
            k: v for k, v in row.items() if k not in ("email", "name") and v
        }
        recipients.append(Recipient(email=email, name=name, extra=extra or None))
    return recipients


def _normalize_header(header: object) -> str:
    return _cell_value(header).lower()


def load_recipients_csv(path: Path) -> list[Recipient]:
    rows: list[dict[str, str]] = []
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("CSV file has no header row")
        for row in reader:
            rows.append({k: (v or "") for k, v in row.items() if k})
    return _rows_to_recipients(rows)


def load_recipients_excel(path: Path, sheet: str | int | None = None) -> list[Recipient]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is None:
            ws = wb.active
        elif isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[sheet]

        row_values = list(ws.iter_rows(values_only=True))
        if not row_values:
            raise ValueError("Excel sheet is empty")

        headers = [_normalize_header(h) for h in row_values[0]]
        if not any(headers):
            raise ValueError("Excel file has no header row")

        rows: list[dict[str, str]] = []
        for values in row_values[1:]:
            if not values or all(v is None or str(v).strip() == "" for v in values):
                continue
            row = {
                headers[i]: _cell_value(values[i]) if i < len(values) else ""
                for i in range(len(headers))
                if headers[i]
            }
            rows.append(row)
    finally:
        wb.close()

    return _rows_to_recipients(rows)


def load_recipients_json(path: Path) -> list[Recipient]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("JSON file must contain a list of recipient objects")

    recipients: list[Recipient] = []
    for item in data:
        if isinstance(item, str):
            recipients.append(Recipient(email=item.strip()))
            continue
        if not isinstance(item, dict) or "email" not in item:
            raise ValueError("Each JSON item needs an 'email' field or be a string")
        recipients.append(
            Recipient(
                email=str(item["email"]).strip(),
                name=str(item.get("name", "")).strip(),
                extra={
                    str(k): str(v)
                    for k, v in item.items()
                    if k not in ("email", "name")
                } or None,
            )
        )
    return recipients


def load_recipients(path: Path, sheet: str | int | None = None) -> list[Recipient]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_recipients_csv(path)
    if suffix == ".json":
        return load_recipients_json(path)
    if suffix in (".xlsx", ".xlsm"):
        return load_recipients_excel(path, sheet=sheet)
    raise ValueError("Recipients file must be .xlsx, .csv, or .json")


def load_recipients_upload(
    content: bytes,
    filename: str,
    sheet: str | int | None = None,
) -> list[Recipient]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ValueError("CSV file has no header row")
        rows = [{k: (v or "") for k, v in row.items() if k} for row in reader]
        return _rows_to_recipients(rows)
    if suffix == ".json":
        data = json.loads(content.decode("utf-8"))
        if not isinstance(data, list):
            raise ValueError("JSON file must contain a list of recipient objects")
        recipients: list[Recipient] = []
        for item in data:
            if isinstance(item, str):
                recipients.append(Recipient(email=item.strip()))
                continue
            if not isinstance(item, dict) or "email" not in item:
                raise ValueError("Each JSON item needs an 'email' field or be a string")
            recipients.append(
                Recipient(
                    email=str(item["email"]).strip(),
                    name=str(item.get("name", "")).strip(),
                    extra={
                        str(k): str(v)
                        for k, v in item.items()
                        if k not in ("email", "name")
                    } or None,
                )
            )
        return recipients
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            if sheet is None:
                ws = wb.active
            elif isinstance(sheet, int):
                ws = wb.worksheets[sheet]
            else:
                ws = wb[sheet]

            row_values = list(ws.iter_rows(values_only=True))
            if not row_values:
                raise ValueError("Excel sheet is empty")

            headers = [_normalize_header(h) for h in row_values[0]]
            if not any(headers):
                raise ValueError("Excel file has no header row")

            rows: list[dict[str, str]] = []
            for values in row_values[1:]:
                if not values or all(v is None or str(v).strip() == "" for v in values):
                    continue
                row = {
                    headers[i]: _cell_value(values[i]) if i < len(values) else ""
                    for i in range(len(headers))
                    if headers[i]
                }
                rows.append(row)
        finally:
            wb.close()
        return _rows_to_recipients(rows)
    raise ValueError("Upload an Excel (.xlsx), CSV, or JSON file")


def render_template(text: str, recipient: Recipient, defaults: dict[str, str]) -> str:
    context = dict(defaults)
    context["email"] = recipient.email
    context["name"] = recipient.name or recipient.email.split("@")[0]
    if recipient.extra:
        context.update(recipient.extra)
    return Template(text).safe_substitute(context)


def build_message(
    config: SmtpConfig,
    recipient: Recipient,
    subject: str,
    body_text: str,
    body_html: str | None,
    reply_to: str | None,
) -> MIMEMultipart:
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["To"] = (
        f"{recipient.name} <{recipient.email}>" if recipient.name else recipient.email
    )

    if config.from_name:
        msg["From"] = f"{config.from_name} <{config.from_email}>"
    else:
        msg["From"] = config.from_email or config.username

    if reply_to:
        msg["Reply-To"] = reply_to

    msg.attach(MIMEText(body_text, "plain", "utf-8"))
    if body_html:
        msg.attach(MIMEText(body_html, "html", "utf-8"))

    return msg


def connect_smtp(config: SmtpConfig) -> smtplib.SMTP:
    if config.use_tls:
        server = smtplib.SMTP(config.host, config.port, timeout=30)
        server.ehlo()
        server.starttls(context=ssl.create_default_context())
        server.ehlo()
    else:
        server = smtplib.SMTP_SSL(config.host, config.port, timeout=30)

    server.login(config.username, config.password)
    return server


def send_bulk(
    recipients: Iterable[Recipient],
    subject_template: str,
    body_template: str,
    *,
    html_template: str | None = None,
    config: SmtpConfig | None = None,
    reply_to: str | None = None,
    delay_seconds: float = 1.0,
    dry_run: bool = False,
    template_defaults: dict[str, str] | None = None,
    on_progress: Callable[[int, int, str, bool, str | None], None] | None = None,
) -> tuple[int, int]:
    """Send emails to all recipients. Returns (success_count, failure_count)."""
    defaults = template_defaults or {}
    reply_to = reply_to or os.getenv("SMTP_REPLY_TO", "").strip() or None

    recipient_list = list(recipients)
    if not recipient_list:
        log.warning("No recipients to send to.")
        return 0, 0

    log.info("Preparing to send to %d recipient(s)", len(recipient_list))

    if dry_run:
        for r in recipient_list:
            subj = render_template(subject_template, r, defaults)
            log.info("[DRY RUN] Would send to %s — subject: %s", r.email, subj)
        return len(recipient_list), 0

    config = config or SmtpConfig.from_env()

    server = connect_smtp(config)
    success = 0
    failed = 0

    try:
        for i, recipient in enumerate(recipient_list, start=1):
            subject = render_template(subject_template, recipient, defaults)
            body_text = render_template(body_template, recipient, defaults)
            body_html = (
                render_template(html_template, recipient, defaults)
                if html_template
                else None
            )

            msg = build_message(
                config, recipient, subject, body_text, body_html, reply_to
            )

            try:
                server.sendmail(
                    config.from_email or config.username,
                    recipient.email,
                    msg.as_string(),
                )
                success += 1
                log.info("[%d/%d] Sent to %s", i, len(recipient_list), recipient.email)
                if on_progress:
                    on_progress(i, len(recipient_list), recipient.email, True, None)
            except smtplib.SMTPException as exc:
                failed += 1
                log.error("[%d/%d] Failed for %s: %s", i, len(recipient_list), recipient.email, exc)
                if on_progress:
                    on_progress(i, len(recipient_list), recipient.email, False, str(exc))

            if i < len(recipient_list) and delay_seconds > 0:
                time.sleep(delay_seconds)
    finally:
        server.quit()

    log.info("Done. Sent: %d, Failed: %d", success, failed)
    return success, failed


def read_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Send one email campaign to multiple recipients at once."
    )
    parser.add_argument(
        "-r",
        "--recipients",
        type=Path,
        default=Path("recipients.xlsx"),
        help="Excel, CSV, or JSON file with recipient list (default: recipients.xlsx)",
    )
    parser.add_argument(
        "--sheet",
        default=os.getenv("RECIPIENTS_SHEET"),
        help="Excel sheet name or 0-based index (default: first sheet)",
    )
    parser.add_argument(
        "-s",
        "--subject",
        type=str,
        help="Email subject (supports $name, $email, and Excel column placeholders)",
    )
    parser.add_argument(
        "--subject-file",
        type=Path,
        help="Read subject from a text file",
    )
    parser.add_argument(
        "-b",
        "--body",
        type=str,
        help="Plain-text email body",
    )
    parser.add_argument(
        "--body-file",
        type=Path,
        default=Path("email_body.txt"),
        help="Plain-text body file (default: email_body.txt)",
    )
    parser.add_argument(
        "--html-file",
        type=Path,
        help="Optional HTML body file",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=float(os.getenv("SEND_DELAY_SECONDS", "1")),
        help="Seconds between sends (default: 1)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview sends without connecting to SMTP",
    )
    return parser.parse_args(argv)


def resolve_recipients_path(path: Path) -> Path:
    if path.exists():
        return path

    if path.name in ("recipients.xlsx", "recipients.csv"):
        alternate = Path("recipients.csv" if path.suffix.lower() == ".xlsx" else "recipients.xlsx")
        if alternate.exists():
            log.info("Using %s", alternate)
            return alternate

    return path


def parse_sheet_arg(sheet: str | None) -> str | int | None:
    if sheet is None or sheet.strip() == "":
        return None
    sheet = sheet.strip()
    if sheet.isdigit():
        return int(sheet)
    return sheet


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    recipients_path = resolve_recipients_path(args.recipients)

    if not recipients_path.exists():
        log.error("Recipients file not found: %s", recipients_path)
        log.error("Create recipients.xlsx in Excel with columns: email, name, company, ...")
        return 1

    if args.subject_file:
        subject = read_text_file(args.subject_file).strip()
    elif args.subject:
        subject = args.subject
    else:
        log.error("Provide --subject or --subject-file")
        return 1

    if args.body:
        body = args.body
    elif args.body_file.exists():
        body = read_text_file(args.body_file)
    else:
        log.error("Provide --body or create %s", args.body_file)
        return 1

    html = read_text_file(args.html_file) if args.html_file and args.html_file.exists() else None
    if args.html_file and not args.html_file.exists():
        log.error("HTML file not found: %s", args.html_file)
        return 1

    try:
        recipients = load_recipients(recipients_path, sheet=parse_sheet_arg(args.sheet))
    except (ValueError, json.JSONDecodeError) as exc:
        log.error("%s", exc)
        return 1

    config: SmtpConfig | None = None
    if not args.dry_run:
        try:
            config = SmtpConfig.from_env()
        except ValueError as exc:
            log.error("%s", exc)
            return 1

    success, failed = send_bulk(
        recipients,
        subject,
        body,
        html_template=html,
        config=config,
        delay_seconds=args.delay,
        dry_run=args.dry_run,
    )

    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
